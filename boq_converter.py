"""
NRE Group – Universal BOQ PDF → Excel Converter
General-purpose, multi-strategy extraction with professional Excel output.
"""

import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import unicodedata
from collections import Counter

# ─── Page Config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NRE BOQ Converter",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;800&family=DM+Sans:wght@300;400;500&display=swap');

  html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

  .main-header {
    background: linear-gradient(135deg, #0f2340 0%, #1a3a6b 60%, #0d4f8c 100%);
    padding: 2rem 2.5rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
    box-shadow: 0 8px 32px rgba(0,0,0,0.25);
  }
  .main-header h1 {
    font-family: 'Syne', sans-serif;
    font-weight: 800;
    font-size: 2rem;
    color: #ffffff;
    margin: 0 0 0.25rem 0;
    letter-spacing: -0.5px;
  }
  .main-header p {
    color: #a8c4e0;
    font-size: 0.95rem;
    margin: 0;
    font-weight: 300;
  }

  .stat-card {
    background: #f8faff;
    border: 1px solid #dde8f5;
    border-left: 4px solid #1a6bbf;
    border-radius: 8px;
    padding: 1rem 1.25rem;
    margin-bottom: 0.5rem;
  }
  .stat-card .label { color: #5a7898; font-size: 0.78rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; }
  .stat-card .value { font-family: 'Syne', sans-serif; font-size: 1.6rem; font-weight: 700; color: #0f2340; line-height: 1.1; }

  .section-badge {
    display: inline-block;
    background: #e8f1fb;
    color: #1a5fa8;
    font-size: 0.72rem;
    font-weight: 600;
    padding: 0.15rem 0.55rem;
    border-radius: 999px;
    letter-spacing: 0.3px;
    text-transform: uppercase;
    margin-bottom: 0.35rem;
  }

  .stDownloadButton > button {
    background: linear-gradient(135deg, #1a6bbf, #0d4f8c) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.65rem 1.5rem !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.92rem !important;
    box-shadow: 0 4px 14px rgba(26,107,191,0.35) !important;
    transition: all 0.2s ease !important;
  }
  .stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(26,107,191,0.45) !important;
  }

  .warning-box {
    background: #fff8e1;
    border: 1px solid #f0c040;
    border-left: 4px solid #f0a500;
    border-radius: 8px;
    padding: 0.85rem 1.1rem;
    font-size: 0.88rem;
    color: #7a5500;
    margin-bottom: 1rem;
  }
  .success-box {
    background: #eaf7f0;
    border: 1px solid #b2dfc8;
    border-left: 4px solid #27a263;
    border-radius: 8px;
    padding: 0.85rem 1.1rem;
    font-size: 0.88rem;
    color: #1a5c3a;
    margin-bottom: 1rem;
  }
  div[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

# ─── Header ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>📋 NRE Group – BOQ Converter</h1>
  <p>Universal Bill of Quantities · PDF → Professional Excel · Multi-Strategy Extraction</p>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def clean_text(text: str) -> str:
    """Normalize Unicode, collapse whitespace, strip."""
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", str(text))
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def is_numeric_cell(val: str) -> bool:
    """Return True if the cell looks like a number/amount."""
    cleaned = re.sub(r"[,\s\u00a0]", "", val)
    return bool(re.fullmatch(r"-?\d+(\.\d+)?", cleaned))


def parse_number(val: str) -> float | None:
    """Parse a number string, returning float or None."""
    cleaned = re.sub(r"[,\s\u00a0]", "", str(val))
    try:
        return float(cleaned)
    except ValueError:
        return None


def is_header_row(row: list[str]) -> bool:
    """Heuristic: row contains typical BOQ header keywords."""
    joined = " ".join(row).upper()
    keywords = {"ITEM", "DESCRIPTION", "UNIT", "QTY", "QUANTITY",
                "RATE", "AMOUNT", "NO", "REF", "TOTAL"}
    hits = sum(1 for kw in keywords if kw in joined)
    return hits >= 2


def is_skip_row(row: list[str]) -> bool:
    """True for blank / page-header / page-footer rows."""
    non_empty = [c for c in row if c.strip()]
    if not non_empty:
        return True
    joined = " ".join(non_empty).upper()
    skip_phrases = [
        "PAGE", "CONTINUED", "CARRY FORWARD", "BROUGHT FORWARD",
        "SUBTOTAL", "SUB-TOTAL", "B/F", "C/F",
    ]
    for phrase in skip_phrases:
        if phrase in joined:
            return True
    return False


# ══════════════════════════════════════════════════════════════════
#  COLUMN SCHEMA DETECTION
# ══════════════════════════════════════════════════════════════════

SCHEMA_PATTERNS = {
    6: ["Item No.", "Description", "Unit", "Quantity", "Rate", "Amount"],
    5: ["Item No.", "Description", "Unit", "Quantity", "Amount"],
    4: ["Item No.", "Description", "Quantity", "Amount"],
    3: ["Item No.", "Description", "Amount"],
}

def detect_schema(rows: list[list[str]]) -> int:
    """Guess the dominant column count from the extracted rows."""
    lengths = [len(r) for r in rows if len(r) > 1]
    if not lengths:
        return 6
    counter = Counter(lengths)
    dominant = counter.most_common(1)[0][0]
    return min(dominant, 6) if dominant > 6 else dominant


def normalise_rows(rows: list[list[str]], target_cols: int) -> list[list[str]]:
    """Pad or trim every row to exactly target_cols columns."""
    result = []
    for row in rows:
        row = [clean_text(c) for c in row]
        if len(row) > target_cols:
            # Merge excess middle columns into description
            row = [row[0]] + [" ".join(row[1: len(row) - target_cols + 2])] + row[len(row) - target_cols + 2:]
        while len(row) < target_cols:
            row.append("")
        result.append(row[:target_cols])
    return result


# ══════════════════════════════════════════════════════════════════
#  EXTRACTION STRATEGIES
# ══════════════════════════════════════════════════════════════════

STRATEGIES = [
    # (vertical, horizontal, snap_y)
    ("lines", "lines", 3),
    ("lines", "text", 3),
    ("text",  "text", 3),
    ("text",  "text", 5),
    ("explicit_vertical_lines", "text", 3),
]


def _extract_with_strategy(page, v_strat, h_strat, snap_y) -> list[list[str]]:
    tables = page.extract_tables({
        "vertical_strategy": v_strat,
        "horizontal_strategy": h_strat,
        "snap_y_tolerance": snap_y,
        "snap_x_tolerance": 3,
        "join_tolerance": 3,
        "edge_min_length": 10,
        "min_words_vertical": 1,
        "min_words_horizontal": 1,
    })
    rows = []
    for table in (tables or []):
        for row in table:
            if row and any(c for c in row if c and c.strip()):
                rows.append([clean_text(c) for c in row])
    return rows


def _extract_word_fallback(page) -> list[list[str]]:
    """
    Word-cluster fallback: group words by Y-position, then by X-clusters
    to reconstruct rows when no table structure is detected.
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    if not words:
        return []

    # Group by approximate Y (snap 4 px)
    y_groups: dict[int, list] = {}
    for w in words:
        y_key = round(w["top"] / 4) * 4
        y_groups.setdefault(y_key, []).append(w)

    rows = []
    for y_key in sorted(y_groups):
        line_words = sorted(y_groups[y_key], key=lambda w: w["x0"])
        row_text = [w["text"] for w in line_words]
        rows.append(row_text)

    return rows


def extract_boq_data(
    pdf_file,
    strategy_index: int = 0,
    min_cols: int = 2,
    progress_cb=None,
) -> tuple[list[list[str]], int]:
    """
    Extract BOQ rows from PDF using the selected strategy.
    Returns (rows, pages_processed).
    """
    v, h, snap = STRATEGIES[strategy_index]
    all_rows: list[list[str]] = []
    pages_done = 0

    with pdfplumber.open(pdf_file) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            if progress_cb:
                progress_cb(i / total, f"Page {i+1}/{total}")

            rows = _extract_with_strategy(page, v, h, snap)

            # If strategy yields nothing meaningful, try word fallback
            meaningful = [r for r in rows if len(r) >= min_cols]
            if not meaningful:
                rows = _extract_word_fallback(page)

            for row in rows:
                if len(row) < min_cols:
                    continue
                if is_skip_row(row):
                    continue
                if is_header_row(row):
                    continue
                all_rows.append(row)

            pages_done = i + 1

    return all_rows, pages_done


# ══════════════════════════════════════════════════════════════════
#  POST-PROCESSING
# ══════════════════════════════════════════════════════════════════

def classify_row_type(row: list[str], n_cols: int) -> str:
    """Classify each row as 'section', 'item', 'total', or 'data'."""
    if n_cols < 2:
        return "data"
    item_col = row[0].strip()
    desc_col = row[1].strip() if n_cols > 1 else ""
    last_col = row[-1].strip()
    second_last = row[-2].strip() if n_cols > 1 else ""

    upper_desc = desc_col.upper()
    if any(kw in upper_desc for kw in ["TOTAL", "GRAND TOTAL", "PRO-SUM", "PROSUM", "SUM"]):
        return "total"
    if not item_col and not is_numeric_cell(last_col) and desc_col:
        return "section"
    if re.match(r"^[A-Z]?\d+(\.\d+)*$", item_col):
        return "item"
    return "data"


def auto_detect_sections(df: pd.DataFrame) -> pd.DataFrame:
    """
    Walk through rows and assign a 'Section' group wherever a section-heading
    row is detected.
    """
    sections = []
    current_section = "General"
    for _, row in df.iterrows():
        rtype = row.get("_type", "data")
        if rtype == "section":
            current_section = row.iloc[1] if len(row) > 1 else "—"
        sections.append(current_section)
    df = df.copy()
    df.insert(0, "Section", sections)
    return df


def compute_summary(df: pd.DataFrame, amount_col: str) -> pd.DataFrame:
    """Build a per-section summary from the data."""
    if "_type" not in df.columns or amount_col not in df.columns:
        return pd.DataFrame()

    df = df.copy()
    df["_amount_num"] = df[amount_col].apply(parse_number)
    summary = (
        df[df["_type"] == "item"]
        .groupby("Section", sort=False)["_amount_num"]
        .sum()
        .reset_index()
    )
    summary.columns = ["Section", "Sub-Total Amount"]
    grand = summary["Sub-Total Amount"].sum()
    summary = pd.concat([
        summary,
        pd.DataFrame([{"Section": "GRAND TOTAL", "Sub-Total Amount": grand}])
    ], ignore_index=True)
    return summary


# ══════════════════════════════════════════════════════════════════
#  EXCEL BUILDER  (openpyxl for full formatting control)
# ══════════════════════════════════════════════════════════════════

def build_excel(df_clean: pd.DataFrame, df_summary: pd.DataFrame, project_name: str) -> bytes:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Colours ──────────────────────────────────────────────────
    C_HEADER_BG  = "1F4E78"
    C_HEADER_FG  = "FFFFFF"
    C_SECTION_BG = "D6E4F0"
    C_SECTION_FG = "0F2340"
    C_TOTAL_BG   = "FFF2CC"
    C_TOTAL_FG   = "7A5500"
    C_GRAND_BG   = "C6EFCE"
    C_GRAND_FG   = "1A5C3A"
    C_ALT_ROW    = "F0F6FC"
    C_BORDER     = "B0C4D8"

    thin  = Side(style="thin",   color=C_BORDER)
    thick = Side(style="medium", color="1F4E78")

    def hdr_border():
        return Border(top=thick, bottom=thick, left=thin, right=thin)

    def cell_border():
        return Border(top=thin, bottom=thin, left=thin, right=thin)

    def apply_header(ws, row_idx, headers):
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=h)
            cell.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
            cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = hdr_border()

    # ─────────────────────────────────────────────────────────────
    #  Sheet 1 – Detailed BOQ
    # ─────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Detailed BOQ"

    # Title banner
    display_cols = [c for c in df_clean.columns if not c.startswith("_")]
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_cols))
    title_cell = ws1.cell(row=1, column=1,
                          value=f"BILL OF QUANTITIES – {project_name.upper()}")
    title_cell.font      = Font(name="Arial", bold=True, size=13, color=C_HEADER_FG)
    title_cell.fill      = PatternFill("solid", fgColor="0F2340")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    apply_header(ws1, 2, display_cols)
    ws1.row_dimensions[2].height = 22

    # Detect amount column index for number formatting
    amount_col_letters = []
    num_fmt = '#,##0.00'

    data_rows = df_clean[display_cols].values.tolist()
    types     = df_clean["_type"].tolist() if "_type" in df_clean.columns else ["data"] * len(data_rows)

    for r_idx, (row_data, rtype) in enumerate(zip(data_rows, types), start=3):
        ws1.row_dimensions[r_idx].height = 15
        is_alt = (r_idx % 2 == 0)

        for c_idx, val in enumerate(row_data, 1):
            col_name = display_cols[c_idx - 1]
            cell = ws1.cell(row=r_idx, column=c_idx)
            cell.border = cell_border()

            # Try numeric parse for rate/qty/amount columns
            num_val = parse_number(str(val)) if val not in (None, "") else None
            if num_val is not None and any(
                kw in col_name.upper() for kw in ["RATE", "AMOUNT", "QTY", "QUANTITY", "TOTAL"]
            ):
                cell.value       = num_val
                cell.number_format = num_fmt
                cell.alignment   = Alignment(horizontal="right", vertical="center")
                amount_col_letters.append(get_column_letter(c_idx))
            else:
                cell.value     = val if val not in (None, "") else ""
                cell.alignment = Alignment(
                    horizontal="left" if c_idx <= 2 else "center",
                    vertical="center", wrap_text=(c_idx == 2)
                )

            # Row colouring
            if rtype == "section":
                cell.font = Font(name="Arial", bold=True, size=9.5, color=C_SECTION_FG)
                cell.fill = PatternFill("solid", fgColor=C_SECTION_BG)
            elif rtype == "total":
                cell.font = Font(name="Arial", bold=True, size=9.5, color=C_TOTAL_FG)
                cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
            else:
                cell.font = Font(name="Arial", size=9)
                if is_alt:
                    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)

    # Column widths
    col_widths = []
    for c_idx, col_name in enumerate(display_cols, 1):
        upper = col_name.upper()
        if "DESC" in upper:
            w = 55
        elif "SECTION" in upper:
            w = 28
        elif "ITEM" in upper or "NO" in upper:
            w = 10
        elif "UNIT" in upper:
            w = 9
        elif any(k in upper for k in ["RATE", "AMOUNT", "QTY", "QUANTITY"]):
            w = 16
        else:
            w = 14
        ws1.column_dimensions[get_column_letter(c_idx)].width = w

    ws1.freeze_panes = "A3"

    # ─────────────────────────────────────────────────────────────
    #  Sheet 2 – Grand Summary  (auto-calculated)
    # ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Grand Summary")

    ws2.merge_cells("A1:C1")
    t2 = ws2.cell(row=1, column=1, value=f"GRAND SUMMARY – {project_name.upper()}")
    t2.font      = Font(name="Arial", bold=True, size=13, color=C_HEADER_FG)
    t2.fill      = PatternFill("solid", fgColor="0F2340")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    apply_header(ws2, 2, ["No.", "Section / Description", "Sub-Total Amount (Rs.)"])
    ws2.row_dimensions[2].height = 22
    ws2.column_dimensions["A"].width = 7
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 24

    if not df_summary.empty:
        total_rows = len(df_summary) - 1   # last row is Grand Total
        data_start = 3
        for i, (_, srow) in enumerate(df_summary.iterrows(), start=data_start):
            is_grand = str(srow.iloc[0]).upper() == "GRAND TOTAL"
            ws2.row_dimensions[i].height = 16

            no_cell   = ws2.cell(row=i, column=1, value="" if is_grand else str(i - data_start + 1))
            desc_cell = ws2.cell(row=i, column=2, value=srow.iloc[0])
            amt_cell  = ws2.cell(row=i, column=3, value=srow.iloc[1])

            for cell in (no_cell, desc_cell, amt_cell):
                cell.border = cell_border()
                cell.alignment = Alignment(vertical="center")

            amt_cell.number_format = num_fmt
            amt_cell.alignment = Alignment(horizontal="right", vertical="center")

            if is_grand:
                for cell in (no_cell, desc_cell, amt_cell):
                    cell.font = Font(name="Arial", bold=True, size=10, color=C_GRAND_FG)
                    cell.fill = PatternFill("solid", fgColor=C_GRAND_BG)
            elif i % 2 == 0:
                for cell in (no_cell, desc_cell, amt_cell):
                    cell.font = Font(name="Arial", size=10)
                    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)
            else:
                for cell in (no_cell, desc_cell, amt_cell):
                    cell.font = Font(name="Arial", size=10)

    ws2.freeze_panes = "A3"

    # ─────────────────────────────────────────────────────────────
    #  Sheet 3 – Extraction Log
    # ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Extraction Log")
    ws3.merge_cells("A1:B1")
    log_title = ws3.cell(row=1, column=1, value="EXTRACTION METADATA")
    log_title.font = Font(name="Arial", bold=True, size=11, color=C_HEADER_FG)
    log_title.fill = PatternFill("solid", fgColor="0F2340")
    log_title.alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 40

    from datetime import datetime
    meta = [
        ("Project Name", project_name),
        ("Total Rows Extracted", len(df_clean)),
        ("Columns Detected", ", ".join(display_cols)),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Tool", "NRE Group BOQ Converter v2.0"),
    ]
    for r_idx, (k, v) in enumerate(meta, start=2):
        ws3.cell(row=r_idx, column=1, value=k).font = Font(name="Arial", bold=True, size=9.5)
        ws3.cell(row=r_idx, column=2, value=str(v)).font = Font(name="Arial", size=9.5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════
#  SIDEBAR – Settings
# ══════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### ⚙️ Extraction Settings")
    strategy_label = {
        0: "Lines + Lines (best for bordered tables)",
        1: "Lines + Text (mixed)",
        2: "Text + Text (general, recommended)",
        3: "Text + Text (relaxed tolerance)",
        4: "Explicit lines + Text (CAD/engineering PDFs)",
    }
    strategy_choice = st.selectbox(
        "PDF Strategy",
        options=list(strategy_label.keys()),
        format_func=lambda x: strategy_label[x],
        index=2,
        help="If extraction looks wrong, try a different strategy."
    )
    min_cols = st.slider("Minimum columns per row", 2, 5, 2)
    project_name = st.text_input("Project Name", value="BOQ Project", max_chars=80)

    st.markdown("---")
    st.markdown("### 📐 Column Mapping")
    st.caption("Override the auto-detected column names if needed:")
    custom_col_names = st.text_area(
        "Column names (one per line)",
        value="Item No.\nDescription\nUnit\nQuantity\nRate\nAmount",
        height=160,
        help="Leave blank to auto-detect from PDF."
    )
    override_cols = [c.strip() for c in custom_col_names.strip().splitlines() if c.strip()]

    st.markdown("---")
    st.markdown("### 🗑️ Row Filtering")
    min_non_empty = st.slider("Min non-empty cells per row", 1, 5, 2,
        help="Rows with fewer filled cells than this are discarded.")


# ══════════════════════════════════════════════════════════════════
#  MAIN AREA
# ══════════════════════════════════════════════════════════════════

uploaded_file = st.file_uploader(
    "Drop your BOQ PDF here",
    type="pdf",
    help="Any BOQ PDF – bordered tables, text-based, or CAD-exported."
)

if not uploaded_file:
    st.markdown("""
    <div style="text-align:center;padding:3rem 1rem;color:#7a9bbc;">
      <div style="font-size:3.5rem;margin-bottom:0.75rem;">📄</div>
      <div style="font-size:1.1rem;font-weight:600;color:#1a3a6b;">Upload a BOQ PDF to begin</div>
      <div style="font-size:0.85rem;margin-top:0.4rem;">Supports any format: bordered tables, text-only, scanned (text layer required)</div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Run Extraction ─────────────────────────────────────────────────────────
progress_bar = st.progress(0, text="Initialising extraction…")

def update_progress(frac, msg):
    progress_bar.progress(frac, text=msg)

with st.spinner(""):
    raw_rows, pages_done = extract_boq_data(
        uploaded_file,
        strategy_index=strategy_choice,
        min_cols=min_cols,
        progress_cb=update_progress,
    )

progress_bar.progress(1.0, text="Extraction complete ✓")

# ── Filter thin rows ───────────────────────────────────────────────────────
filtered_rows = [
    r for r in raw_rows
    if sum(1 for c in r if c.strip()) >= min_non_empty
]

# ── Detect schema & normalise ──────────────────────────────────────────────
n_cols = detect_schema(filtered_rows)
norm_rows = normalise_rows(filtered_rows, n_cols)

# Apply column names
if override_cols and len(override_cols) == n_cols:
    col_names = override_cols
elif override_cols and len(override_cols) != n_cols:
    st.markdown(f"""<div class="warning-box">⚠️ You provided {len(override_cols)} column names
    but {n_cols} columns were detected. Auto-naming will be used.</div>""", unsafe_allow_html=True)
    col_names = SCHEMA_PATTERNS.get(n_cols, [f"Col {i+1}" for i in range(n_cols)])
else:
    col_names = SCHEMA_PATTERNS.get(n_cols, [f"Col {i+1}" for i in range(n_cols)])

# Build DataFrame
df = pd.DataFrame(norm_rows, columns=col_names)
df["_type"] = [classify_row_type(r, n_cols) for r in norm_rows]

# Detect amount column
amount_col_candidates = [c for c in col_names if any(
    k in c.upper() for k in ["AMOUNT", "TOTAL", "RATE"]
)]
amount_col = amount_col_candidates[-1] if amount_col_candidates else col_names[-1]

# Auto section detection (if section column not already present)
df = auto_detect_sections(df)

# Summary
df_summary = compute_summary(df, amount_col)

# ── Display stats ──────────────────────────────────────────────────────────
st.markdown(f"""<div class="success-box">
  ✅ Extracted <strong>{len(df)}</strong> rows from <strong>{pages_done}</strong> pages using strategy
  "<em>{strategy_label[strategy_choice]}</em>"
</div>""", unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown(f'<div class="stat-card"><div class="label">Total Rows</div><div class="value">{len(df)}</div></div>', unsafe_allow_html=True)
with col2:
    item_count = int(df[df["_type"] == "item"].shape[0]) if "_type" in df.columns else "—"
    st.markdown(f'<div class="stat-card"><div class="label">Line Items</div><div class="value">{item_count}</div></div>', unsafe_allow_html=True)
with col3:
    section_count = int(df[df["_type"] == "section"].shape[0]) if "_type" in df.columns else "—"
    st.markdown(f'<div class="stat-card"><div class="label">Sections</div><div class="value">{section_count}</div></div>', unsafe_allow_html=True)
with col4:
    st.markdown(f'<div class="stat-card"><div class="label">Columns</div><div class="value">{n_cols}</div></div>', unsafe_allow_html=True)

# ── Preview ────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📋 Detailed BOQ", "📊 Grand Summary"])

display_df = df[[c for c in df.columns if not c.startswith("_")]].copy()

with tab1:
    st.dataframe(display_df, use_container_width=True, height=420)

with tab2:
    if not df_summary.empty:
        st.dataframe(df_summary, use_container_width=True)
        grand_val = df_summary.iloc[-1, 1]
        if pd.notna(grand_val):
            st.markdown(f"**Grand Total: Rs. {grand_val:,.2f}**")
    else:
        st.info("Summary could not be generated automatically. Ensure the Amount column contains numeric values.")

# ── Build & Download Excel ─────────────────────────────────────────────────
excel_bytes = build_excel(df, df_summary, project_name)
safe_name   = re.sub(r"[^\w\-]", "_", project_name)

st.download_button(
    label="📥 Download Professional Excel",
    data=excel_bytes,
    file_name=f"{safe_name}_BOQ.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=False,
)

st.markdown("---")
st.caption("NRE Group BOQ Converter v2.0 · Universal multi-strategy extraction · openpyxl professional output")
